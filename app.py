"""
app.py

Giao diện web Flask cho hệ thống đo tôm.
File này phục vụ UI, API cấu hình, upload input, chạy pipeline bằng subprocess,
đọc kết quả JSON, export CSV/XLSX và hiệu chuẩn SCALE. Mọi đọc/ghi settings.json
đều đi qua settings_loader hoặc module config/size cùng cấp thư mục dự án.
"""

import csv
import io
import json
import mimetypes
import os
import re
import subprocess
import sys
import threading
import time
from pathlib import Path
from typing import Any

from flask import Flask, Response, jsonify, render_template, request, send_from_directory
from openpyxl import Workbook, load_workbook

from config import load_config_values
from settings_loader import pull_setting_warnings, read_setting, save_setting
from size import load_size_values


BASE_DIR = Path(__file__).resolve().parent

mimetypes.add_type("text/css; charset=utf-8", ".css")
mimetypes.add_type("application/javascript; charset=utf-8", ".js")

app = Flask(
    __name__,
    template_folder=str(BASE_DIR / "templates"),
    static_folder=str(BASE_DIR / "static"),
    static_url_path="/static",
)
app.config["MAX_CONTENT_LENGTH"] = 4 * 1024 * 1024 * 1024
app.json.ensure_ascii = False

# Settings bridge

CONFIG_KEYS = [
    "INPUT_DIR", "OUTPUT_DIR", "CLEAR_OUTPUT", "CLEAR_INPUT", "CHUNK_MODE",
    "SCALE", "CONF_DET", "CONF_SEG", "BBOX_PAD", "TOUCH_THRESHOLD",
    "TARGET_FPS", "CONVEYOR_VERTICAL", "SAVE",
]
BOOL_KEYS = {"CLEAR_OUTPUT", "CLEAR_INPUT", "CHUNK_MODE", "CONVEYOR_VERTICAL", "SAVE"}

RESULT_COLS = ["run", "source_file", "track_id", "frame_idx", "pixel_length", "real_length_mm", "size"]
MAX_SCALE_BYTE = 4 * 1024 * 1024
CSV_MIMETYPE = "text/csv"
XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
EXPORT_FILENAME_PREFIX = "shrimp"
SCALE_FILE_HEADER = "mm"
SCALE_FILE_COLUMN_ERROR = "File scale phải có đúng 1 cột với header mm"
SCALE_FILE_HEADER_ERROR = "Header file scale phải là mm"
SCALE_FILE_EMPTY_ERROR = "Không tìm thấy giá trị mm hợp lệ trong file"


def _get_config() -> dict:
    """
    Lấy config mới nhất cho API Flask.

    Hàm gọi config.load_config_values(), nên mọi lỗi hoặc thiếu settings.json
    đều được settings_loader phục hồi và ghi default nếu cần.
    """
    return load_config_values()


def _save_config(values: dict) -> None:
    """
    Lưu config qua settings_loader.
    Giữ IMG_EXTS, VID_EXTS và key người dùng thêm tay trong section config.
    """
    current = read_setting().get("config", {})
    if not isinstance(current, dict):
        current = {}
    merged  = {**current, **{k: values[k] for k in CONFIG_KEYS if k in values}}
    save_setting("config", merged)


def _get_sizes() -> dict:
    """
    Lấy bảng kích cỡ mới nhất cho UI.

    Trả về ranges dưới dạng list để JSON response dễ dùng ở frontend, trong khi
    size.load_size_values() vẫn dùng tuple float cho logic backend.
    """
    values = load_size_values()
    return {
        "ranges": {k: list(v) for k, v in values["SIZE_RANGES"].items()},
        "undersize_label": values["UNDERSIZE_LABEL"],
        "oversize_label":  values["OVERSIZE_LABEL"],
        "fallback_label":  values["FALLBACK_LABEL"],
    }


# Kiểm tra payload

def _validate_config(raw: dict) -> dict:
    """
    Kiểm tra và chuẩn hóa payload config từ client.

    Chỉ các key trong CONFIG_KEYS được nhận. Hàm ép kiểu bool, float, int và
    ném ValueError với thông báo có thể hiển thị trực tiếp lên UI.
    """
    current = _get_config()
    data    = {**current, **{k: raw[k] for k in CONFIG_KEYS if k in raw}}

    for k in ["INPUT_DIR", "OUTPUT_DIR"]:
        v = str(data.get(k, "")).strip()
        if not v:
            raise ValueError(f"{k} không được để trống")
        data[k] = v

    for k in BOOL_KEYS:
        v = data[k]
        data[k] = (v.strip().lower() in {"1", "true", "yes", "on"}
                   if isinstance(v, str) else bool(v))

    for k, lo, hi in [
        ("SCALE",           1e-5, None),
        ("CONF_DET",        0.0,  1.0),
        ("CONF_SEG",        0.0,  1.0),
        ("TOUCH_THRESHOLD", 0.0,  None),
        ("TARGET_FPS",      0.0,  None),
    ]:
        try:
            v = float(data[k])
        except Exception:
            raise ValueError(f"{k} phải là số thực")
        if lo is not None and v < lo:
            raise ValueError(f"{k} phải >= {lo}")
        if hi is not None and v > hi:
            raise ValueError(f"{k} phải <= {hi}")
        data[k] = v

    try:
        v = int(data["BBOX_PAD"])
        if v < 0:
            raise ValueError
        data["BBOX_PAD"] = v
    except (ValueError, TypeError):
        raise ValueError("BBOX_PAD phải là số nguyên không âm")

    return data


def _validate_sizes(raw: dict) -> dict:
    """
    Kiểm tra và chuẩn hóa payload bảng kích cỡ từ client.

    Ranges phải là object nhãn -> [từ, đến], không âm, không rỗng và không
    chồng lấp. Nhãn ngoại cỡ rỗng sẽ được thay bằng nhãn mặc định.
    """
    ranges_raw = raw.get("ranges", {})
    if not isinstance(ranges_raw, dict):
        raise ValueError("ranges phải là object")

    result: list[tuple[str, float, float]] = []
    labels: set[str] = set()

    for lbl, bounds in ranges_raw.items():
        lbl = str(lbl).strip()
        if not lbl:
            raise ValueError("Tên cỡ không được để trống")
        if lbl in labels:
            raise ValueError(f"Cỡ '{lbl}' bị trùng")
        labels.add(lbl)
        if not isinstance(bounds, (list, tuple)) or len(bounds) != 2:
            raise ValueError(f"Cỡ '{lbl}' cần [từ, đến]")
        try:
            lo, hi = float(bounds[0]), float(bounds[1])
        except Exception:
            raise ValueError(f"Cỡ '{lbl}': giá trị phải là số")
        if lo < 0 or hi < 0:
            raise ValueError(f"Cỡ '{lbl}': không được âm")
        if lo >= hi:
            raise ValueError(f"Cỡ '{lbl}': mốc đầu phải nhỏ hơn mốc cuối")
        result.append((lbl, lo, hi))

    result.sort(key=lambda x: x[1])
    for i in range(1, len(result)):
        a, _, a_hi = result[i - 1]
        b, b_lo, _ = result[i]
        if b_lo < a_hi:
            raise ValueError(f"Cỡ '{a}' và '{b}' bị chồng lấp")

    return {
        "ranges":          {lbl: [lo, hi] for lbl, lo, hi in result},
        "undersize_label": str(raw.get("undersize_label", "")).strip() or "Ngoại cỡ nhỏ",
        "oversize_label":  str(raw.get("oversize_label",  "")).strip() or "Ngoại cỡ lớn",
        "fallback_label":  str(raw.get("fallback_label",  "")).strip() or "Ngoại cỡ",
    }


# Đường dẫn

def _abs(path: str) -> Path:
    """Đổi path tương đối theo BASE_DIR thành đường dẫn tuyệt đối."""
    p = Path(path)
    return p.resolve() if p.is_absolute() else (BASE_DIR / p).resolve()


def _input_dir() -> Path:
    """Trả về INPUT_DIR tuyệt đối và tạo thư mục nếu thiếu."""
    p = _abs(_get_config()["INPUT_DIR"])
    p.mkdir(parents=True, exist_ok=True)
    return p


def _output_dir() -> Path:
    """Trả về OUTPUT_DIR tuyệt đối và tạo thư mục nếu thiếu."""
    p = _abs(_get_config()["OUTPUT_DIR"])
    p.mkdir(parents=True, exist_ok=True)
    return p


def _log_path() -> Path:
    """Trả về đường dẫn pipeline.log trong OUTPUT_DIR."""
    return _output_dir() / "pipeline.log"


# File input

def _allowed_ext() -> set[str]:
    """Trả về tập đuôi file input được phép upload."""
    cfg = _get_config()
    return {e.lower() for e in [*cfg["IMG_EXTS"], *cfg["VID_EXTS"]]}


def _safe_name(filename: str) -> str:
    """Chuẩn hóa tên file upload để không chứa path hoặc ký tự nguy hiểm."""
    name = Path(str(filename).replace("\\", "/")).name.strip()
    name = re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip(" .")
    return name or f"upload_{time.time_ns()}"


def _unique_dest(directory: Path, name: str) -> Path:
    """Tạo đường dẫn không trùng trong directory bằng hậu tố _1, _2 nếu cần."""
    p = directory / name
    stem, suffix, i = Path(name).stem, Path(name).suffix, 1
    while p.exists():
        p = directory / f"{stem}_{i}{suffix}"
        i += 1
    return p


def _file_info(p: Path) -> dict:
    """Trả thông tin file input cho UI."""
    s = p.stat()
    return {"name": p.name, "size": s.st_size, "mtime": s.st_mtime, "suffix": p.suffix.lower()}


# Trạng thái pipeline

_lock:    threading.Lock       = threading.Lock()
_proc:    subprocess.Popen | None = None
_running: bool                 = False
_t_start: float | None        = None
_t_end:   float | None        = None
_retcode: int | None          = None   # 0 = thành công, != 0 = lỗi


def _watch_pipeline(proc: subprocess.Popen) -> None:
    """Chạy trong thread riêng, chờ subprocess kết thúc và cập nhật trạng thái."""
    global _running, _t_end, _retcode
    code = proc.wait()
    with _lock:
        _retcode = code
        _running = False
        _t_end   = time.time()


def _status() -> dict:
    """Trả trạng thái subprocess pipeline hiện tại cho API status."""
    return {
        "running":    _running,
        "returncode": None if _running else _retcode,
        "started_at": _t_start,
        "ended_at":   None if _running else _t_end,
    }


# Tiện ích kết quả

def _run_dirs() -> list[Path]:
    """Liệt kê các thư mục run trong OUTPUT_DIR, mới nhất đứng trước."""
    d = _output_dir()
    return sorted(
        [p for p in d.iterdir() if p.is_dir()],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )


def _json_files(run_dir: Path) -> list[Path]:
    """Liệt kê các file kết quả JSON của một run."""
    return sorted(run_dir.glob("*/*_results.json"))


def _selected_run_dir(run_name: str | None) -> Path | None:
    """Chọn run theo tên, hoặc run mới nhất nếu không truyền run_name."""
    runs = _run_dirs()
    if run_name:
        return next((r for r in runs if r.name == run_name), None)
    return runs[0] if runs else None


def _read_result_json(path: Path) -> dict[str, Any] | None:
    """Đọc một file kết quả JSON, trả None nếu file hỏng hoặc không phải object."""
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeDecodeError, json.JSONDecodeError):
        return None
    return data if isinstance(data, dict) else None


def _image_url(raw: str | None) -> str | None:
    """
    Đổi đường dẫn ảnh debug thành URL /outputs an toàn cho UI.

    Chỉ trả URL nếu ảnh nằm trong OUTPUT_DIR và file đang tồn tại.
    """
    if not raw:
        return None
    root = _output_dir().resolve()
    p    = Path(raw)
    p    = p.resolve() if p.is_absolute() else (BASE_DIR / p).resolve()
    try:
        rel = p.relative_to(root)
    except ValueError:
        return None
    return f"/outputs/{rel.as_posix()}" if p.exists() else None


def _norm_images(images: dict | None) -> dict:
    """Chuẩn hóa dict ảnh debug trong JSON kết quả thành URL frontend dùng được."""
    if not images:
        return {}
    return {
        k: [u for x in v if (u := _image_url(x))] if isinstance(v, list) else _image_url(v)
        for k, v in images.items()
    }


def _results_for_run(run_name: str | None) -> dict:
    """
    Gom toàn bộ kết quả của một run thành response cho UI.

    Hàm bỏ qua file JSON hỏng, chuẩn hóa đường dẫn ảnh debug và trả shape
    `{"run": name, "sources": [...]}`. Nếu chưa có run thì trả run None.
    """
    sel = _selected_run_dir(run_name)
    if sel is None:
        return {"run": None, "sources": []}

    sources = []
    for jf in _json_files(sel):
        d = _read_result_json(jf)
        if d is None:
            continue
        sources.append({
            "source_file":     d.get("source_file",    jf.parent.name),
            "source_stem":     d.get("source_stem",    jf.parent.name),
            "processed_at":    d.get("processed_at"),
            "scale_mm_per_px": d.get("scale_mm_per_px"),
            "shrimps": [
                dict(s, images=_norm_images(s.get("images")))
                for s in d.get("shrimps", [])
            ],
        })
    return {"run": sel.name, "sources": sources}


def _export_rows(data: dict) -> list[list]:
    """Chuyển dữ liệu results thành các dòng dùng chung cho CSV và XLSX."""
    rows = [RESULT_COLS]
    for src in data["sources"]:
        for s in src["shrimps"]:
            rows.append([
                data["run"], src["source_file"],
                s.get("track_id"), s.get("frame_idx"),
                s.get("pixel_length"), s.get("real_length_mm"), s.get("size"),
            ])
    return rows


def _export_filename(run_name: str | None, suffix: str) -> str:
    """Tạo tên file export an toàn theo run và suffix."""
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(run_name or "empty"))
    return f"{EXPORT_FILENAME_PREFIX}_{safe_name}.{suffix}"


# Excel

def _xlsx_bytes(rows: list[list]) -> bytes:
    """Tạo workbook XLSX trong bộ nhớ từ danh sách dòng."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# File scale import

def _decode_bytes(raw: bytes) -> str:
    """Decode file text upload bằng các encoding thường gặp của CSV tiếng Việt."""
    for enc in ("utf-8-sig", "utf-8", "cp1258", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _parse_positive(cell: Any) -> float | None:
    """Parse một ô thành số dương, trả None nếu ô không hợp lệ."""
    text = str(cell or "").strip().replace(",", ".")
    if not re.fullmatch(r"\d+(?:\.\d+)?|\.\d+", text):
        return None
    n = float(text)
    return n if n > 0 else None


def _is_blank_cell(cell: Any) -> bool:
    """Kiểm tra một cell CSV/XLSX có đang trống sau khi strip không."""
    return cell is None or str(cell).strip() == ""


def _single_scale_cell(row: list[Any] | tuple[Any, ...], row_number: int, header: bool = False) -> Any:
    """Lấy ô duy nhất của một dòng file scale, ném lỗi nếu có nhiều cột."""
    non_empty_indexes = [i for i, cell in enumerate(row) if not _is_blank_cell(cell)]
    if not non_empty_indexes:
        if header:
            raise ValueError(SCALE_FILE_HEADER_ERROR)
        raise ValueError(f"Dòng {row_number}: giá trị mm không hợp lệ")
    if len(non_empty_indexes) != 1 or non_empty_indexes[0] != 0:
        raise ValueError(SCALE_FILE_COLUMN_ERROR)
    return row[0]


def _read_scale_rows(rows: list[list[Any]] | list[tuple[Any, ...]]) -> list[float]:
    """
    Đọc các dòng file scale đã tách cell.

    Dòng đầu phải là header mm, các dòng sau phải là số dương. Hàm ném
    ValueError với thông báo rõ để endpoint trả lỗi trực tiếp cho UI.
    """
    if not rows:
        raise ValueError(SCALE_FILE_EMPTY_ERROR)

    header = _single_scale_cell(rows[0], 1, header=True)
    if str(header).strip().lower() != SCALE_FILE_HEADER:
        raise ValueError(SCALE_FILE_HEADER_ERROR)

    values = []
    for row_number, row in enumerate(rows[1:], start=2):
        cell = _single_scale_cell(row, row_number)
        number = _parse_positive(cell)
        if number is None:
            raise ValueError(f"Dòng {row_number}: giá trị mm không hợp lệ")
        values.append(number)

    if not values:
        raise ValueError(SCALE_FILE_EMPTY_ERROR)
    return values


def _read_col_csv(raw: bytes) -> list[float]:
    """Đọc file scale CSV bắt buộc có 1 cột header mm."""
    text = _decode_bytes(raw)
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    return _read_scale_rows(list(csv.reader(io.StringIO(text), dialect)))


def _read_col_xlsx(raw: bytes) -> list[float]:
    """Đọc file scale XLSX bắt buộc có 1 cột header mm."""
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError("File XLSX không hợp lệ") from e
    try:
        ws = wb.active
        if ws is None:
            raise ValueError("Không tìm thấy sheet trong XLSX")
        return _read_scale_rows(list(ws.iter_rows(values_only=True)))
    finally:
        wb.close()


# Routes

@app.get("/")
def index():
    """Render trang UI chính."""
    return render_template("index.html")


# File input

@app.get("/api/files/input")
def list_input():
    """Trả danh sách file hiện có trong INPUT_DIR."""
    d = _input_dir()
    return jsonify({
        "files": [_file_info(p) for p in sorted(d.iterdir()) if p.is_file()]
    })


@app.post("/api/files/upload")
def upload():
    """
    Nhận nhiều file upload và lưu vào INPUT_DIR.

    File có đuôi không nằm trong IMG_EXTS hoặc VID_EXTS bị đưa vào danh sách
    rejected. File trùng tên được tự thêm hậu tố để không ghi đè.
    """
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "Chưa chọn file"}), 400
    allowed  = _allowed_ext()
    d        = _input_dir()
    saved, rejected = [], []
    for f in files:
        ext = Path(f.filename or "").suffix.lower()
        if ext not in allowed:
            rejected.append({"name": f.filename, "reason": "Định dạng không hỗ trợ"})
            continue
        dest = _unique_dest(d, _safe_name(f.filename))
        f.save(dest)
        saved.append(_file_info(dest))
    return jsonify({"saved": saved, "rejected": rejected})


@app.delete("/api/files/input/<path:filename>")
def delete_input(filename: str):
    """Xóa một file input nếu đường dẫn vẫn nằm trong INPUT_DIR."""
    d = _input_dir().resolve()
    t = (d / filename).resolve()
    try:
        t.relative_to(d)
    except ValueError:
        return jsonify({"error": "Tên file không hợp lệ"}), 400
    if t.is_file():
        t.unlink()
    return jsonify({"ok": True})


# Pipeline

@app.post("/api/pipeline/run")
def run_pipeline():
    """
    Khởi động pipeline bằng subprocess main.py.

    Trước khi chạy, app ép config và size được load qua settings_loader để
    settings.json thiếu hoặc hỏng được phục hồi. Response gồm trạng thái ban
    đầu và warnings để UI hiển thị toast nếu vừa phục hồi settings.json.
    """
    global _proc, _running, _t_start, _t_end, _retcode
    warnings: list[str] = []
    with _lock:
        if _running:
            return jsonify({"error": "Pipeline đang chạy"}), 409
        _get_config()
        _get_sizes()
        warnings = pull_setting_warnings()
        _output_dir().mkdir(parents=True, exist_ok=True)
        _log_path().write_text("", encoding="utf-8")
        _t_start, _t_end, _retcode = time.time(), None, None
        _running = True
        _proc    = subprocess.Popen(
            [sys.executable, str(BASE_DIR / "main.py")],
            cwd=str(BASE_DIR),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.STDOUT,
        )
        threading.Thread(
            target=_watch_pipeline,
            args=(_proc,),
            daemon=True,
            name="pipeline-watcher",
        ).start()
    return jsonify({"ok": True, "status": _status(), "warnings": warnings})


@app.get("/api/pipeline/status")
def pipeline_status():
    """Trả trạng thái chạy của subprocess pipeline."""
    return jsonify(_status())


@app.get("/api/pipeline/log")
def pipeline_log():
    """
    Đọc log pipeline theo offset.

    UI gọi endpoint này để tail log mà không tải lại toàn bộ file. Nếu offset
    lớn hơn kích thước hiện tại, endpoint reset offset về 0.
    """
    try:
        offset = max(0, int(request.args.get("offset", "0")))
    except ValueError:
        offset = 0
    p = _log_path()
    if not p.exists():
        return jsonify({"content": "", "offset": 0, "size": 0})
    size = p.stat().st_size
    if offset > size:
        offset = 0
    with p.open("rb") as f:
        f.seek(offset)
        data = f.read()
        nxt  = f.tell()
    return jsonify({
        "content": data.decode("utf-8", errors="replace"),
        "offset":  nxt,
        "size":    size,
    })


# Cấu hình

@app.get("/api/config")
def get_config():
    """Trả config hiện tại và warnings settings nếu có."""
    return jsonify({
        "config": _get_config(),
        "warnings": pull_setting_warnings(),
    })


@app.put("/api/config")
def put_config():
    """Validate payload config, lưu vào settings.json và trả config mới."""
    try:
        data = _validate_config(request.get_json(force=True, silent=True) or {})
        _save_config(data)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({
        "config": _get_config(),
        "warnings": pull_setting_warnings(),
    })


# Kích cỡ

@app.get("/api/config/sizes")
def get_sizes():
    """Trả bảng phân loại kích cỡ hiện tại và warnings settings nếu có."""
    return jsonify({
        "sizes": _get_sizes(),
        "warnings": pull_setting_warnings(),
    })


@app.put("/api/config/sizes")
def put_sizes():
    """Validate bảng size từ UI, lưu section size và trả bảng mới."""
    try:
        data = _validate_sizes(request.get_json(force=True, silent=True) or {})
        save_setting("size", {
            "SIZE_RANGES":     data["ranges"],
            "UNDERSIZE_LABEL": data["undersize_label"],
            "OVERSIZE_LABEL":  data["oversize_label"],
            "FALLBACK_LABEL":  data["fallback_label"],
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({
        "sizes": _get_sizes(),
        "warnings": pull_setting_warnings(),
    })


# Kết quả

@app.get("/api/results/runs")
def result_runs():
    """Trả danh sách run kết quả cùng số tôm đã ghi trong JSON."""
    runs = []
    for rd in _run_dirs():
        count = 0
        for jf in _json_files(rd):
            d = _read_result_json(jf)
            if d is not None:
                count += len(d.get("shrimps", []))
        runs.append({
            "name":         rd.name,
            "mtime":        rd.stat().st_mtime,
            "shrimp_count": count,
        })
    return jsonify({"runs": runs})


@app.get("/api/results")
def results():
    """Trả kết quả chi tiết của run được chọn hoặc run mới nhất."""
    return jsonify(_results_for_run(request.args.get("run")))


@app.get("/api/results/export-csv")
def export_csv():
    """Xuất kết quả của run được chọn ra CSV."""
    data = _results_for_run(request.args.get("run"))
    buf  = io.StringIO()
    csv.writer(buf).writerows(_export_rows(data))
    return Response(
        buf.getvalue(),
        mimetype=CSV_MIMETYPE,
        headers={"Content-Disposition": f"attachment; filename={_export_filename(data['run'], 'csv')}"},
    )


@app.get("/api/results/export-excel")
def export_excel():
    """Xuất kết quả của run được chọn ra XLSX."""
    data = _results_for_run(request.args.get("run"))
    return Response(
        _xlsx_bytes(_export_rows(data)),
        mimetype=XLSX_MIMETYPE,
        headers={"Content-Disposition": f"attachment; filename={_export_filename(data['run'], 'xlsx')}"},
    )


# Hiệu chuẩn scale

@app.post("/api/calibrate/import-measurements")
def import_measurements():
    """
    Nhập file đo thực tế để điền nhanh bảng hiệu chuẩn SCALE.

    File CSV hoặc XLSX phải có đúng một cột, header `mm`, các dòng sau là số
    dương. Values được ghép tuần tự với danh sách tôm frontend gửi lên. Endpoint
    trả 400 nếu file sai cấu trúc hoặc payload run/rows không hợp lệ.
    """
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "Chưa chọn file"}), 400
    run_name = str(request.form.get("run") or "").strip()
    if not run_name:
        return jsonify({"error": "Chưa chọn run"}), 400
    try:
        ordered = json.loads(request.form.get("rows") or "[]")
    except Exception:
        return jsonify({"error": "Danh sách tôm không hợp lệ"}), 400
    if not isinstance(ordered, list) or not ordered:
        return jsonify({"error": "Danh sách tôm trống"}), 400

    raw = f.read(MAX_SCALE_BYTE + 1)
    if len(raw) > MAX_SCALE_BYTE:
        return jsonify({"error": "File vượt quá 4 MB"}), 400

    suffix = Path(f.filename).suffix.lower()
    try:
        if suffix == ".csv":
            mm_values = _read_col_csv(raw)
        elif suffix == ".xlsx":
            mm_values = _read_col_xlsx(raw)
        else:
            return jsonify({"error": "Chỉ hỗ trợ CSV hoặc XLSX"}), 400
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Không đọc được file: {e}"}), 400

    if not mm_values:
        return jsonify({"error": SCALE_FILE_EMPTY_ERROR}), 400

    measurements = [
        {
            "source_file":    str(row.get("source_file", "")),
            "source_stem":    str(row.get("source_stem", "")),
            "track_id":       str(row.get("track_id",    "")),
            "real_length_mm": mm,
        }
        for row, mm in zip(ordered, mm_values)
        if isinstance(row, dict)
    ]
    warnings = []
    if len(mm_values) != len(ordered):
        warnings.append(
            f"File có {len(mm_values)} giá trị, danh sách có {len(ordered)} tôm. "
            f"Đã điền {len(measurements)} dòng."
        )
    return jsonify({
        "measurements":   measurements,
        "count":          len(measurements),
        "expected_count": len(ordered),
        "warnings":       warnings,
    })


@app.post("/api/calibrate")
def calibrate():
    """
    Tính và lưu SCALE mới từ dữ liệu đo thực tế.

    Endpoint lấy pixel_length từ JSON kết quả của run đã chọn, ghép với
    real_length_mm do UI gửi lên, rồi fit hồi quy tuyến tính qua gốc tọa độ:
    real_length_mm = SCALE x pixel_length. SCALE mới được lưu qua _save_config().
    """
    payload  = request.get_json(force=True, silent=True) or {}
    run_name = str(payload.get("run") or "").strip()
    meas     = payload.get("measurements")
    if not run_name:
        return jsonify({"error": "Chưa chọn run"}), 400
    if not isinstance(meas, list) or not meas:
        return jsonify({"error": "Chưa có dữ liệu đo thực tế"}), 400

    run_dir = _selected_run_dir(run_name)
    if run_dir is None:
        return jsonify({"error": f"Không tìm thấy run '{run_name}'"}), 404

    # Bảng tra pixel_length từ file JSON kết quả của backend
    px_index: dict[tuple[str, str], float] = {}
    for jf in _json_files(run_dir):
        d = _read_result_json(jf)
        if d is None:
            continue
        stem = str(d.get("source_stem") or jf.parent.name)
        for s in d.get("shrimps", []):
            px_index[(stem, str(s.get("track_id")))] = float(s.get("pixel_length", 0))

    samples, errors = [], []
    for item in meas:
        if not isinstance(item, dict):
            continue
        stem = str(item.get("source_stem") or "").strip()
        tid  = str(item.get("track_id")    or "").strip()
        try:
            real_mm = float(item["real_length_mm"])
        except Exception:
            errors.append(f"{stem} ID {tid}: mm không hợp lệ")
            continue
        if real_mm <= 0:
            errors.append(f"{stem} ID {tid}: mm phải > 0")
            continue
        px = px_index.get((stem, tid))
        if px is None or px <= 0:
            errors.append(f"{stem} ID {tid}: không tìm thấy pixel_length hợp lệ")
            continue
        samples.append({
            "source_stem":    stem,
            "track_id":       tid,
            "pixel_length":   px,
            "real_length_mm": real_mm,
        })

    if not samples:
        return jsonify({"error": "Không có mẫu hợp lệ", "errors": errors}), 400

    sum_xx = sum(s["pixel_length"] ** 2 for s in samples)
    sum_xy = sum(s["pixel_length"] * s["real_length_mm"] for s in samples)
    if sum_xx == 0:
        return jsonify({"error": "pixel_length bằng 0"}), 400

    scale = sum_xy / sum_xx
    rmse  = (
        sum((s["pixel_length"] * scale - s["real_length_mm"]) ** 2 for s in samples)
        / len(samples)
    ) ** 0.5

    try:
        data = _validate_config({"SCALE": round(scale, 6)})
        _save_config(data)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    return jsonify({
        "scale":   round(scale, 6),
        "rmse_mm": round(rmse,  6),
        "count":   len(samples),
        "errors":  errors,
        "config":  _get_config(),
    })


# Phục vụ ảnh debug

@app.get("/outputs/<path:filename>")
def output_file(filename: str):
    """Phục vụ ảnh debug trong OUTPUT_DIR và chặn path traversal."""
    root   = _output_dir().resolve()
    target = (root / filename).resolve()
    try:
        target.relative_to(root)
    except ValueError:
        return jsonify({"error": "Đường dẫn không hợp lệ"}), 400
    return send_from_directory(root, filename)


# Khởi động

if __name__ == "__main__":
    host  = os.environ.get("HOST", "127.0.0.1")
    port  = int(os.environ.get("PORT", "3000"))
    debug = os.environ.get("FLASK_DEBUG", "1").strip().lower() in {"1", "true", "yes", "on"}
    print(f"Shrimp Measure UI: http://{host}:{port}", flush=True)
    app.run(host=host, port=port, debug=debug, use_reloader=False)
